import queue
import logging
import time
import customtkinter as ctk
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Импорты модулей системы 
from network.udp_receiver import UDPReceiver
from processing.imu_processor import IMUProcessor
from processing.filter_manager import FilterManager
from ui.plot_widget import PlotWidget

logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] (%(threadName)s) %(message)s')

class VOnTR_App(ctk.CTk):
    def __init__(self):
        super().__init__()

        # --- Конфигурация главного окна ---
        self.title("ВОнТР v_1 — Визуальная Одометрия Realme")
        self.geometry("1200x800")
        self.minsize(1000, 600)
        ctk.set_appearance_mode("System")
        ctk.set_default_color_theme("blue")

        # --- Инициализация сетевой архитектуры и ядра ---
        self.data_queue = queue.Queue(maxsize=5000)
        self.receiver = None
        self.processor = IMUProcessor()
        self.filter_manager = FilterManager()
        self.export_data = []  # накопление данных для экспорта

        self.is_gathering = False
        self.packet_count = 0
        self._queue_timer_id = None

        self._create_widgets()

    def _set_controls_state(self, state: str):
        """Включение/выключение элементов управления."""
        self.cb_raw.configure(state=state)
        self.cb_exp.configure(state=state)
        self.cb_iir.configure(state=state)
        self.cb_kalman.configure(state=state)

        self.rb_3d.configure(state=state)
        self.rb_x.configure(state=state)
        self.rb_y.configure(state=state)
        self.rb_z.configure(state=state)

        self.btn_start.configure(state=state)
        self.btn_export.configure(state=state)

    def connect_socket(self):
        """Валидация параметров сети и подключение."""
        ip = self.ip_entry.get().strip()
        port_str = self.port_entry.get().strip()

        if not ip or not port_str:
            self.status_label.configure(text="Статус: Заполните IP и Порт!", text_color="red")
            return

        try:
            port = int(port_str)
            if not (1024 <= port <= 65535):
                raise ValueError()
        except ValueError:
            self.status_label.configure(text="Статус: Некорректный порт!", text_color="red")
            return

        self.status_label.configure(text="Статус: Подключение успешно", text_color="green")
        self._set_controls_state("normal")
        self.btn_connect.configure(text="Отключить сокет", fg_color="gray", hover_color="#555555", command=self.disconnect_socket)

        self.ip_entry.configure(state="disabled")
        self.port_entry.configure(state="disabled")

    def disconnect_socket(self):
        """Отключение сокета и блокировка интерфейса."""
        if self.is_gathering:
            self.stop_gathering()

        self.status_label.configure(text="Статус: Требуется подключение", text_color="#1F6AA5")
        self._set_controls_state("disabled")
        self.btn_reset.configure(state="disabled")

        self.btn_connect.configure(text="Подключить сокет", fg_color="#1F6AA5", hover_color="#144871", command=self.connect_socket)
        self.ip_entry.configure(state="normal")
        self.port_entry.configure(state="normal")

    def check_queue(self):
        """Обработка очереди пакетов и конвейер фильтрации."""
        if not self.is_gathering:
            return

        packets_processed = 0

        while not self.data_queue.empty():
            try:
                packet = self.data_queue.get_nowait()
            except queue.Empty:
                break

            try:
                if not isinstance(packet, dict):
                    continue

                acc_data = packet.get("accel", {})
                if isinstance(acc_data, dict):
                    accel_raw = [float(acc_data.get('x', 0)), float(acc_data.get('y', 0)), float(acc_data.get('z', 0))]
                else:
                    accel_raw = [float(x) for x in acc_data]

                gyro_data = packet.get("gyro", {})
                if isinstance(gyro_data, dict):
                    gyro_raw = [float(gyro_data.get('x', 0)), float(gyro_data.get('y', 0)), float(gyro_data.get('z', 0))]
                else:
                    gyro_raw = [float(x) for x in gyro_data]

                q_data = packet.get("quat", {"w": 1.0, "x": 0.0, "y": 0.0, "z": 0.0})
                if isinstance(q_data, dict):
                    quat = [float(q_data.get('x', 0)), float(q_data.get('y', 0)), float(q_data.get('z', 0)), float(q_data.get('w', 1))]
                else:
                    quat = [float(x) for x in q_data]

                normalized_packet = {
                    "accel": accel_raw,
                    "gyro": gyro_raw,
                    "quat": quat,
                    "timestamp": packet.get("timestamp", time.time())
                }

                raw_data = self.processor.process_packet(normalized_packet)

                if raw_data:
                    filtered_data = self.filter_manager.apply_filters(raw_data)
                    self.plot_widget.add_data_point(filtered_data)

                    self.export_data.append({
                        "packet_id": self.packet_count,
                        "timestamp": packet.get("timestamp", time.time()),
                        "RAW": filtered_data["RAW"],
                        "EXP": filtered_data["EXP"],
                        "IIR": filtered_data["IIR"],
                        "KALMAN": filtered_data["KALMAN"]
                    })

                    packets_processed += 1
                    self.packet_count += 1

            except Exception as e:
                print(f"СБОЙ КОНВЕЙЕРА В MAIN: {e}")
            finally:
                self.data_queue.task_done()

        if packets_processed > 0:
            self.counter_label.configure(text=f"Принято пакетов: {self.packet_count}")
            self.plot_widget.update_plots()

        if self.is_gathering:
            self._queue_timer_id = self.after(33, self.check_queue)

    def export_to_excel(self):
        """Экспорт накопленных данных в Excel с генерацией 2D-графиков по столбцам (в ряд справа)."""
        if not self.export_data:
            self.status_label.configure(text="Статус: Нет данных для экспорта!", text_color="red")
            return

        self.status_label.configure(text="Статус: Экспорт данных...", text_color="orange")
        self.update()

        try:
            from openpyxl.chart import LineChart, Reference
            from openpyxl.chart.series import SeriesLabel
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Данные одометрии"

            font_title = Font(name="Segoe UI", size=14, bold=True, color="1F6AA5")
            font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
            fill_header = PatternFill(start_color="1F6AA5", end_color="1F6AA5", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin', color='E0E0E0'), 
                right=Side(style='thin', color='E0E0E0'), 
                top=Side(style='thin', color='E0E0E0'),
                bottom=Side(style='thin', color='E0E0E0')
            )

            # Главный заголовок
            ws.merge_cells("A1:M1")
            ws["A1"] = "АНАЛИТИЧЕСКИЙ ОТЧЕТ СИСТЕМЫ ФИЛЬТРАЦИИ ИНС ВОнТР"
            ws["A1"].font = font_title
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions.height = 30

            # Структура верхней шапки таблицы
            headers = [
                ("A3:A4", "№ Точки"),
                ("B3:D3", "Сырые данные (RAW)"),
                ("E3:G3", "Экспоненциальный (EXP)"),
                ("H3:J3", "БИХ-фильтр (IIR)"),
                ("K3:M3", "Фильтр Калмана (KALMAN)")
            ]

            for cell_range, text in headers:
                ws.merge_cells(cell_range)
                for row in ws[cell_range]:
                    for cell in row:
                        cell.fill = fill_header
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
                # Фикс: берем координату первой ячейки
                first_cell_coord = cell_range.split(":")[0]
                ws[first_cell_coord].value = text
                ws[first_cell_coord].font = font_header

            # Подзаголовки осей пространства (Строка 4)
            sub_headers = ["X", "Y", "Z"]
            for f_idx in range(4):
                for axis_idx, axis_name in enumerate(sub_headers):
                    col_num = 2 + (f_idx * 3) + axis_idx
                    cell = ws.cell(row=4, column=col_num, value=axis_name)
                    cell.font = font_header
                    cell.fill = fill_header
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            ws["A4"].border = thin_border

            for col_idx in range(1, 14):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = 12

            # Заполнение данными
            for row_idx, data_point in enumerate(self.export_data, start=5):
                p_cell = ws.cell(row=row_idx, column=1, value=data_point["packet_id"])
                p_cell.alignment = Alignment(horizontal="center")
                p_cell.border = thin_border
                
                for f_idx, key in enumerate(["RAW", "EXP", "IIR", "KALMAN"]):
                    coords = data_point[key]
                    for axis_idx in range(3):
                        col_num = 2 + (f_idx * 3) + axis_idx
                        cell = ws.cell(row=row_idx, column=col_num, value=float(coords[axis_idx]))
                        cell.border = thin_border
                        cell.number_format = '0.0000' 

            max_row = len(self.export_data) + 4
            filter_names = ["RAW", "EXP", "IIR", "KALMAN"]
            graph_columns = ["O3", "Y3", "AI3"]

            # Генерация графиков по столбцам (справа от таблицы)
            for axis_idx, axis_name in enumerate(["X", "Y", "Z"]):
                chart = LineChart()
                chart.title = f"Сравнение фильтрации — Ось {axis_name}"
                chart.style = 13  
                chart.y_axis.title = f"Позиция {axis_name} (м)"
                chart.x_axis.title = "Время (пакеты)"
                chart.width = 16   
                chart.height = 11  
                
                for f_idx, filter_label in enumerate(filter_names):
                    target_col = 2 + (f_idx * 3) + axis_idx
                    # Ссылка строго со строки 5, отключив titles_from_data
                    data_ref = Reference(ws, min_col=target_col, min_row=5, max_row=max_row)
                    chart.add_data(data_ref, titles_from_data=False)
                    
                    # Фикс: оборачиваем строку в SeriesLabel для прохождения валидации openpyxl
                    chart.series[f_idx].title = SeriesLabel(v=filter_label)
                
                cats_ref = Reference(ws, min_col=1, min_row=5, max_row=max_row)
                chart.set_categories(cats_ref)
                
                target_anchor = graph_columns[axis_idx]
                ws.add_chart(chart, target_anchor)

            filename = f"vontr_export_{int(time.time())}.xlsx"
            wb.save(filename)
            self.status_label.configure(text=f"Успех: Данные и ГРАФИКИ сохранены в {filename}", text_color="green")

        except Exception as e:
            self.status_label.configure(text="Статус: Ошибка экспорта!", text_color="red")
            logging.error(f"Ошибка экспорта Excel с графиками по столбцам: {e}")


    def _create_widgets(self):
        """Создание и компоновка всех элементов интерфейса."""
        self.grid_columnconfigure(0, weight=25, minsize=260)
        self.grid_columnconfigure(1, weight=75)
        self.grid_rowconfigure(0, weight=1)

        # ===================== ЛЕВАЯ ПАНЕЛЬ =====================
        self.sidebar = ctk.CTkFrame(self, corner_radius=0)
        self.sidebar.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)

        self.sidebar.grid_columnconfigure(0, weight=1)
        for i in range(11):
            self.sidebar.grid_rowconfigure(i, weight=0 if i != 5 else 1)

        self.title_label = ctk.CTkLabel(self.sidebar, text="ВОнТР v_1", font=ctk.CTkFont(size=22, weight="bold"))
        self.title_label.grid(row=0, column=0, padx=20, pady=(10, 2), sticky="ew")

        self.subtitle_label = ctk.CTkLabel(self.sidebar, text="Инерциальная одометрия IMU", font=ctk.CTkFont(size=12, slant="italic"))
        self.subtitle_label.grid(row=1, column=0, padx=20, pady=(0, 10), sticky="ew")

        # Настройки UDP
        self.net_frame = ctk.CTkFrame(self.sidebar)
        self.net_frame.grid(row=2, column=0, sticky="ew", padx=15, pady=2)
        self.net_frame.grid_columnconfigure(0, weight=1)

        self.net_title = ctk.CTkLabel(self.net_frame, text="Настройки UDP Connection", font=ctk.CTkFont(weight="bold"))
        self.net_title.grid(row=0, column=0, sticky="w", padx=15, pady=(8, 4))

        self.ip_label = ctk.CTkLabel(self.net_frame, text="IP Адрес:")
        self.ip_label.grid(row=1, column=0, sticky="w", padx=15, pady=(2, 0))
        self.ip_entry = ctk.CTkEntry(self.net_frame, height=26)
        self.ip_entry.insert(0, "127.0.0.1")
        self.ip_entry.grid(row=2, column=0, sticky="ew", padx=15, pady=(0, 4))

        self.port_label = ctk.CTkLabel(self.net_frame, text="Порт:")
        self.port_label.grid(row=3, column=0, sticky="w", padx=15, pady=(2, 0))
        self.port_entry = ctk.CTkEntry(self.net_frame, height=26)
        self.port_entry.insert(0, "5005")
        self.port_entry.grid(row=4, column=0, sticky="ew", padx=15, pady=(0, 6))

        self.btn_connect = ctk.CTkButton(self.net_frame, text="Подключить сокет", fg_color="#1F6AA5", hover_color="#144871", height=28, command=self.connect_socket)
        self.btn_connect.grid(row=5, column=0, sticky="ew", padx=15, pady=(0, 10))

        # Фильтры
        self.filter_frame = ctk.CTkFrame(self.sidebar)
        self.filter_frame.grid(row=3, column=0, sticky="ew", padx=15, pady=2)
        self.filter_frame.grid_columnconfigure(0, weight=1)

        self.filter_title = ctk.CTkLabel(self.filter_frame, text="Алгоритмы фильтрации", font=ctk.CTkFont(weight="bold"))
        self.filter_title.grid(row=0, column=0, sticky="w", padx=15, pady=(8, 4))

        self.cb_raw = ctk.CTkCheckBox(self.filter_frame, text="Сырые данные (RAW)", state="disabled", command=self._on_filter_changed)
        self.cb_raw.select()
        self.cb_raw.grid(row=1, column=0, sticky="w", padx=15, pady=3)

        self.cb_exp = ctk.CTkCheckBox(self.filter_frame, text="Экспоненциальный фильтр", state="disabled", command=self._on_filter_changed)
        self.cb_exp.select()
        self.cb_exp.grid(row=2, column=0, sticky="w", padx=15, pady=3)

        self.cb_iir = ctk.CTkCheckBox(self.filter_frame, text="БИХ-фильтр (IIR)", state="disabled", command=self._on_filter_changed)
        self.cb_iir.select()
        self.cb_iir.grid(row=3, column=0, sticky="w", padx=15, pady=3)

        self.cb_kalman = ctk.CTkCheckBox(self.filter_frame, text="Фильтр Калмана", state="disabled", command=self._on_filter_changed)
        self.cb_kalman.select()
        self.cb_kalman.grid(row=4, column=0, sticky="w", padx=15, pady=(3, 10))

        # Выбор осей
        self.axis_frame = ctk.CTkFrame(self.sidebar)
        self.axis_frame.grid(row=4, column=0, sticky="ew", padx=15, pady=2)
        self.axis_frame.grid_columnconfigure(0, weight=1)

        self.axis_title = ctk.CTkLabel(self.axis_frame, text="Отображаемые оси", font=ctk.CTkFont(weight="bold"))
        self.axis_title.grid(row=0, column=0, sticky="w", padx=15, pady=(8, 4))

        self.axis_var = ctk.StringVar(value="3D")
        self.rb_3d = ctk.CTkRadioButton(self.axis_frame, text="3D Траектория", state="disabled", variable=self.axis_var, value="3D", command=self._on_axis_changed)
        self.rb_3d.grid(row=1, column=0, sticky="w", padx=15, pady=3)
        self.rb_x = ctk.CTkRadioButton(self.axis_frame, text="Ось X (Времени)", state="disabled", variable=self.axis_var, value="X", command=self._on_axis_changed)
        self.rb_x.grid(row=2, column=0, sticky="w", padx=15, pady=3)
        self.rb_y = ctk.CTkRadioButton(self.axis_frame, text="Ось Y (Времени)", state="disabled", variable=self.axis_var, value="Y", command=self._on_axis_changed)
        self.rb_y.grid(row=3, column=0, sticky="w", padx=15, pady=3)
        self.rb_z = ctk.CTkRadioButton(self.axis_frame, text="Ось Z (Времени)", state="disabled", variable=self.axis_var, value="Z", command=self._on_axis_changed)
        self.rb_z.grid(row=4, column=0, sticky="w", padx=15, pady=(3, 10))

        self.spacer = ctk.CTkFrame(self.sidebar, fg_color="transparent")
        self.spacer.grid(row=5, column=0, sticky="nsew")

        # Статус и счётчик
        self.info_frame = ctk.CTkFrame(self.sidebar, fg_color="transparent")
        self.info_frame.grid(row=6, column=0, sticky="ew", padx=15, pady=2)

        self.status_label = ctk.CTkLabel(self.info_frame, text="Статус: Требуется подключение", text_color="#1F6AA5", font=ctk.CTkFont(weight="bold"))
        self.status_label.pack(anchor="w", pady=1)

        self.counter_label = ctk.CTkLabel(self.info_frame, text="Принято пакетов: 0")
        self.counter_label.pack(anchor="w", pady=1)

        # Кнопки
        self.btn_reset = ctk.CTkButton(self.sidebar, text="Сбросить траекторию", fg_color="gray", hover_color="#555555", state="disabled", height=32, command=self.reset_trajectory)
        self.btn_reset.grid(row=7, column=0, sticky="ew", padx=15, pady=3)

        self.btn_export = ctk.CTkButton(self.sidebar, text="Экспорт в Excel", fg_color="#E28743", hover_color="#B86221", state="disabled", height=32, command=self.export_to_excel)
        self.btn_export.grid(row=8, column=0, sticky="ew", padx=15, pady=3)

        self.btn_stop = ctk.CTkButton(self.sidebar, text="ОСТАНОВИТЬ", fg_color="red", hover_color="darkred", state="disabled", height=32, command=self.stop_gathering)
        self.btn_stop.grid(row=9, column=0, sticky="ew", padx=15, pady=3)

        self.btn_start = ctk.CTkButton(self.sidebar, text="ЗАПУСТИТЬ СБОР", fg_color="green", hover_color="darkgreen", state="disabled", height=32, command=self.start_gathering)
        self.btn_start.grid(row=10, column=0, sticky="ew", padx=15, pady=(3, 15))

        # ===================== ПРАВАЯ ПАНЕЛЬ (ГРАФИК) =====================
        self.plot_container = ctk.CTkFrame(self)
        self.plot_container.grid(row=0, column=1, sticky="nsew", padx=10, pady=10)

        self.plot_widget = PlotWidget(self.plot_container)
        self.plot_widget.pack(fill="both", expand=True, padx=5, pady=5)


    def start_gathering(self):
        """Запуск приёма данных."""
        if self._queue_timer_id is not None:
            self.after_cancel(self._queue_timer_id)
            self._queue_timer_id = None

        ip = self.ip_entry.get().strip()
        try:
            port = int(self.port_entry.get().strip())
        except ValueError:
            self.status_label.configure(text="Статус: Неверный порт!", text_color="red")
            return

        self.receiver = UDPReceiver(ip, port, self.data_queue)
        try:
            self.receiver.start()
            self.is_gathering = True

            self.btn_start.configure(state="disabled")
            self.btn_stop.configure(state="normal")
            self.btn_reset.configure(state="disabled")
            self.btn_connect.configure(state="disabled")
            self.btn_export.configure(state="disabled")

            self.status_label.configure(text="Статус: Приём данных...", text_color="green")
            self._queue_timer_id = self.after(33, self.check_queue)  # ~30 FPS

        except Exception as e:
            self.status_label.configure(text="Статус: Ошибка запуска!", text_color="red")
            logging.error(f"Критический сбой инициализации сокета: {e}")

    def stop_gathering(self):
        """Остановка приёма данных."""
        self.is_gathering = False
        if self._queue_timer_id is not None:
            self.after_cancel(self._queue_timer_id)
            self._queue_timer_id = None

        if self.receiver:
            self.receiver.stop()
            self.receiver = None

        self.btn_start.configure(state="normal")
        self.btn_stop.configure(state="disabled")
        self.btn_reset.configure(state="normal")
        self.btn_connect.configure(state="normal")
        self.btn_export.configure(state="normal")

        self.status_label.configure(text="Статус: Подключение успешно", text_color="green")

    def reset_trajectory(self):
        """Сброс всех данных и графиков."""
        self.processor.reset()
        self.filter_manager.reset()
        self.export_data.clear()

        self.plot_widget.clear_plots()
        self.packet_count = 0
        self.counter_label.configure(text="Принято пакетов: 0")

        while not self.data_queue.empty():
            try:
                self.data_queue.get_nowait()
            except queue.Empty:
                break

        self.status_label.configure(text="Статус: Графики очищены", text_color="cyan")

    def _on_filter_changed(self):
        """Обработчик изменения чекбоксов фильтров."""
        filter_mask = {
            "RAW": bool(self.cb_raw.get()),
            "EXP": bool(self.cb_exp.get()),
            "IIR": bool(self.cb_iir.get()),
            "KALMAN": bool(self.cb_kalman.get())
        }
        self.plot_widget.toggle_filters_visibility(filter_mask)

    def _on_axis_changed(self):
        """Обработчик переключения режима отображения."""
        selected_axis = self.axis_var.get()
        self.plot_widget.change_view_mode(selected_axis)